#!/usr/bin/env python3
"""Cruza catalogo Magento CSV con tarifa Excel Miniland (openpyxl)."""
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "CSV" / "Tarifa VENTA EDU SP 2026_01 SCHOOL .xlsx"
CSV_PATH = ROOT / "catalogo-español 2026 CSV.csv"


def parse_csv_all(text: str, delimiter: str) -> list[list[str]]:
    rows: list[list[str]] = []
    row: list[str] = []
    field = ""
    in_quotes = False
    for i, c in enumerate(text):
        if in_quotes:
            if c == '"':
                if i + 1 < len(text) and text[i + 1] == '"':
                    field += '"'
                else:
                    in_quotes = False
            else:
                field += c
        elif c == '"':
            in_quotes = True
        elif c == delimiter:
            row.append(field)
            field = ""
        elif c == "\r":
            continue
        elif c == "\n":
            row.append(field)
            field = ""
            if any(x.strip() for x in row):
                rows.append(row)
            row = []
        else:
            field += c
    row.append(field)
    if any(x.strip() for x in row):
        rows.append(row)
    return rows


def clean_text(s) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    if t.startswith('"') and t.endswith('"'):
        t = t[1:-1].replace('""', '"')
    return t.strip()


def parse_price(val) -> float:
    try:
        return float(clean_text(val).replace(",", "."))
    except ValueError:
        return 0.0


def extract_ref(sku: str, line: str) -> str:
    m = re.search(r"/security/(\d+)_", line)
    if m:
        return m.group(1)
    s = clean_text(sku)
    if s.startswith("5005") and len(s) >= 10:
        return s[-5:]
    return s


def load_tarifa(path: Path) -> dict[str, dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = None
    for i, row in enumerate(rows[:30]):
        cells = [str(c).strip().upper() if c is not None else "" for c in row]
        if any(c == "REF." or c == "REF" for c in cells) and any("PRECIO" in c for c in cells):
            header_row = i
            break
    if header_row is None:
        raise RuntimeError("No se encontró fila cabecera REF/PRECIO en Excel")

    hdr = [str(c).strip().upper() if c is not None else "" for c in rows[header_row]]
    ref_i = next(i for i, h in enumerate(hdr) if h in ("REF.", "REF", "REFERENCIA"))
    precio_i = next(i for i, h in enumerate(hdr) if h == "PRECIO")
    pvpr_i = next((i for i, h in enumerate(hdr) if "PVPR" in h), None)
    desc_i = next((i for i, h in enumerate(hdr) if "DESCRIP" in h), None)

    print(f"Excel cabecera fila {header_row + 1}: REF col {ref_i}, PRECIO col {precio_i}")
    data_start = header_row + 2  # fila subcabecera EURO/PVPR

    tarifa: dict[str, dict] = {}
    for row in rows[data_start :]:
        if not row or ref_i >= len(row):
            continue
        ref_raw = row[ref_i]
        if ref_raw is None:
            continue
        ref = str(int(ref_raw)) if isinstance(ref_raw, (int, float)) else str(ref_raw).strip()
        if not ref or ref == "None":
            continue
        try:
            precio = float(row[precio_i]) if row[precio_i] is not None else 0.0
        except (TypeError, ValueError):
            precio = 0.0
        pvpr = 0.0
        if pvpr_i is not None and pvpr_i < len(row) and row[pvpr_i] is not None:
            try:
                pvpr = float(row[pvpr_i])
            except (TypeError, ValueError):
                pass
        desc = str(row[desc_i]) if desc_i is not None and desc_i < len(row) and row[desc_i] else ""
        tarifa[ref] = {"precio": precio, "pvpr": pvpr, "desc": desc[:60]}

    return tarifa


def load_catalog(path: Path) -> list[dict]:
    products = []
    sku_i, name_i, price_i = 60, 42, 48

    with open(path, encoding="utf-8-sig") as f:
        next(f)  # cabecera Magento
        for raw_line in f:
            line = raw_line.rstrip("\r\n")
            if not line.strip():
                continue
            outer = parse_csv_all(line, ";")
            if not outer:
                continue
            cell = ""
            for c in outer[0]:
                t = clean_text(c)
                if t:
                    cell = t
                    break
            if not cell:
                continue
            inner = parse_csv_all(cell, ",")
            if not inner or not inner[0]:
                continue
            row = inner[0]
            if len(row) <= max(sku_i, name_i, price_i):
                continue
            sku = clean_text(row[sku_i])
            name = clean_text(row[name_i])
            price = parse_price(row[price_i])
            if not name or len(name) > 500:
                continue
            if not re.match(r"^[A-Za-z0-9._-]+$", sku) or len(sku) < 2:
                continue
            ref = extract_ref(sku, cell)
            products.append({"name": name, "sku": sku, "ref": ref, "csv_price": price})
    return products


def main() -> None:
    tarifa = load_tarifa(XLSX)
    print(f"Refs tarifa: {len(tarifa)}, PRECIO>0: {sum(1 for v in tarifa.values() if v['precio'] > 0)}")
    print(f"REF 32171: {tarifa.get('32171')}")

    products = load_catalog(CSV_PATH)
    zero = [p for p in products if p["csv_price"] == 0]
    positive = [p for p in products if p["csv_price"] > 0]

    in_tarifa = [p for p in zero if tarifa.get(p["ref"], {}).get("precio", 0) > 0]
    not_in_tarifa = [p for p in zero if tarifa.get(p["ref"], {}).get("precio", 0) <= 0]

    print(f"\nCatálogo parseado (n8n): {len(products)} productos")
    print(f"CSV price=0: {len(zero)}")
    print(f"CSV price>0: {len(positive)}")
    print(f"Price=0 + PRECIO en Excel: {len(in_tarifa)}  (los del mail)")
    print(f"Price=0 + sin Excel: {len(not_in_tarifa)}  (descatalogados segun Miniland)")

    print("\n--- Price=0 resueltos por Excel (muestra) ---")
    for p in in_tarifa[:15]:
        t = tarifa[p["ref"]]
        print(f"  REF {p['ref']} | PRECIO {t['precio']} | PVPR {t['pvpr']} | {p['name'][:50]}")

    print("\n--- Price=0 sin tarifa (muestra) ---")
    for p in not_in_tarifa[:15]:
        print(f"  REF {p['ref']} | {p['name'][:55]}")

    dolls_zero = [
        p for p in zero
        if re.search(r"doll|muñec|munec|beb[eé]", p["name"], re.I)
    ]
    print(f"\nMuñecos/bebé con price=0: {len(dolls_zero)}")
    print(f"  con tarifa: {sum(1 for p in dolls_zero if tarifa.get(p['ref'], {}).get('precio', 0) > 0)}")
    print(f"  sin tarifa: {sum(1 for p in dolls_zero if tarifa.get(p['ref'], {}).get('precio', 0) <= 0)}")

    # Productos con CSV price>0: ¿es PVP neto o PRECIO?
    sample = [p for p in positive if p["ref"] in tarifa][:5]
    print("\n--- CSV price>0 vs tarifa (muestra) ---")
    for p in sample:
        t = tarifa[p["ref"]]["precio"]
        pvpr = tarifa[p["ref"]]["pvpr"]
        print(
            f"  REF {p['ref']} | CSV {p['csv_price']:.2f} | PRECIO {t:.2f} | "
            f"PVPR/1.21={pvpr/1.21:.2f} | {p['name'][:40]}"
        )


if __name__ == "__main__":
    main()
